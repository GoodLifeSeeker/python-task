import argparse
import csv
import json
import openpyxl
import requests

from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill
from os.path import join, abspath

from constants import URL, CSV_FILE_NAME, HEADERS


def get_csv_data(file):
    """Extracts list of dictionaries from csv file."""
    csv_dicts = []
    with open(file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f, delimiter=";")
        for line in reader:
            csv_dicts.append(line)
    return csv_dicts


def create_excel(data, columns, colored, path):
    """Creates final excel file."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    label_ids_col_idx = None

    # Creating table's headers and checking if the 'labelIds' in it
    for col, column_name in enumerate(columns):
        worksheet.cell(row=1, column=col+1, value=column_name)
        if column_name == 'labelIds':
            label_ids_col_idx = col

    # Filling the table with data returned from server
    for row, item in enumerate(data, start=2):
        for col, column_name in enumerate(columns):
            cell_value = item.get(column_name)
            worksheet.cell(row=row, column=col+1, value=cell_value)

        # Coloring the background if '-c' flag is True
        if colored and 'hu' in item.keys():
            colors = ['#007500', '#FFA500', '#b30000']
            hu_date = datetime.strptime(item['hu'], '%Y-%m-%d')
            now_date = datetime.now()
            diff_in_months = (now_date - hu_date) // timedelta(days=30)

            if diff_in_months < 3:
                fill_color = colors[0][1:]
            elif diff_in_months <= 12:
                fill_color = colors[1][1:]
            else:
                fill_color = colors[2][1:]

            for cell in worksheet[row]:
                cell.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type='solid'
                )
    # Coloring row's font if there's 'labelIds' field and it's not empty
    if label_ids_col_idx:
        for row in worksheet.iter_rows(min_row=2):
            label_ids_value = row[label_ids_col_idx].value
            if label_ids_value:
                font = Font(color=label_ids_value[1:])
                for cell in row:
                    cell.font = font
    # Saving excel file
    workbook.save(path)


# Preparation paths
current_date = datetime.now().date().isoformat()
excel_name = f'vehicles_{current_date}.xlsx'
excel_path = abspath(join('.', excel_name))
csv_file_path = abspath(join('.', CSV_FILE_NAME))

# Getting data from csv file
csv_data = get_csv_data(csv_file_path)

# Sending get request to server
res = requests.post(URL, data=json.dumps(csv_data), headers=HEADERS)
if res.status_code == 200:
    response_json = res.json()
    response_json = sorted(response_json, key=lambda x: x['gruppe'])
else:
    print(f'Request failed with code {res.status_code}: {res.text}')

# Working with input parameters
# Common
parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('-k', '--keys', nargs='+', help='additional keys')
parser.add_argument('-c', '--colored', default=True, help='flag to coloring')
args = parser.parse_args()

# -k/--keys part
k_args = args.keys
excel_columns = ['rnr',]
if k_args is not None:
    excel_columns += k_args
else:
    excel_columns = list(response_json[0].keys())

# -c/--colored part
color_flag = args.colored
if color_flag == 'False':
    color_flag = False
else:
    color_flag = True

# Calling create_excel func
create_excel(response_json, excel_columns, color_flag, excel_path)
